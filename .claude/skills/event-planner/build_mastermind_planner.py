#!/usr/bin/env python3
"""
Build the Lisbon Mastermind 2026 Event Planner — multi-tab .xlsx for Xarlish.

Output: ~/Desktop/lisbon-mastermind-planner-2026.xlsx

Tabs:
  1. Overview         — event metadata + key links
  2. Master Checklist — every task with owner/deadline/status
  3. Budget           — cost line items + vendor + estimated vs actual
  4. Villa Shortlist  — 5-7 villa options with specs
  5. Vendors          — content team, catering, transport, surf, restaurants
  6. Attendees        — guest list + dietary + payment status
  7. Schedule         — 3-day plan with timings
  8. Gift Bag         — swag tracker per item
"""
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ─── Brand style constants ─────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="8A00FF")  # Violet Ray
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Helvetica")
SECTION_FILL = PatternFill("solid", fgColor="1A0033")
SECTION_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FONT = Font(size=20, bold=True, color="8A00FF", name="Helvetica")
SUBTITLE_FONT = Font(size=12, color="4B5563", italic=True)

BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")

# Status colors
DONE_FILL = PatternFill("solid", fgColor="D1FAE5")        # green
PROGRESS_FILL = PatternFill("solid", fgColor="DBEAFE")    # blue
BLOCKED_FILL = PatternFill("solid", fgColor="FECDD3")     # pink
NOT_STARTED_FILL = PatternFill("solid", fgColor="F3F4F6") # grey
ALT_ROW_FILL = PatternFill("solid", fgColor="FAFAFA")

# Payment status
PAID_FILL = PatternFill("solid", fgColor="D1FAE5")
PENDING_FILL = PatternFill("solid", fgColor="FEF3C7")
UNPAID_FILL = PatternFill("solid", fgColor="FECDD3")


# ─── Helpers ───────────────────────────────────────────────────────────
def style_headers(ws, row, cols):
    """Apply violet header style to a row of column letters."""
    for col in cols:
        cell = ws[f"{col}{row}"]
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


def add_status_dropdown(ws, col_letter, first_row, last_row, options):
    dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def apply_status_conditional(ws, col_letter, first_row, last_row, full_range):
    """Color whole row by status column."""
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(formula=[f'${col_letter}{first_row}="Done"'], fill=DONE_FILL),
    )
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(formula=[f'${col_letter}{first_row}="In Progress"'], fill=PROGRESS_FILL),
    )
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(formula=[f'${col_letter}{first_row}="Blocked"'], fill=BLOCKED_FILL),
    )
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(formula=[f'${col_letter}{first_row}="Not Started"'], fill=NOT_STARTED_FILL),
    )


def write_rows(ws, start_row, rows, widths):
    """Write a list of dict-rows to a sheet. Rows must use keys matching widths dict order."""
    for i, row_data in enumerate(rows):
        row = start_row + i
        for col_idx, (key, _) in enumerate(widths.items(), start=1):
            val = row_data.get(key, "")
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        ws.row_dimensions[row].height = 32
    return start_row + len(rows) - 1


def set_widths(ws, widths):
    for col_idx, (_, w) in enumerate(widths.items(), start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ─── Build workbook ────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

STATUS_OPTIONS = ["Not Started", "In Progress", "Blocked", "Done"]
PAYMENT_OPTIONS = ["Unpaid", "Deposit Paid", "Paid in Full"]
VENDOR_STATUS = ["Researching", "Quoted", "Booked", "Paid", "Done"]


# ============= TAB 1: OVERVIEW =============
ov = wb.create_sheet("Overview")
ov.sheet_view.showGridLines = False

ov.merge_cells("A1:C1")
ov["A1"] = "Lisbon Mastermind 2026 — Event Planner"
ov["A1"].font = TITLE_FONT
ov["A1"].alignment = Alignment(horizontal="left", vertical="center")
ov.row_dimensions[1].height = 40

ov.merge_cells("A2:C2")
ov["A2"] = "Trademark RecruiterGTM annual event. First edition. Xarlish leads execution; Reyhan hosts."
ov["A2"].font = SUBTITLE_FONT
ov["A2"].alignment = Alignment(horizontal="left", vertical="center")
ov.row_dimensions[2].height = 24

ov["A4"] = "Field"
ov["B4"] = "Value"
ov["C4"] = "Notes"
for col in ("A4", "B4", "C4"):
    ov[col].fill = HEADER_FILL
    ov[col].font = HEADER_FONT
    ov[col].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ov[col].border = BORDER
ov.row_dimensions[4].height = 28

overview_rows = [
    ("Event name", "Lisbon Mastermind 2026", "First edition, trademark RecruiterGTM event"),
    ("Host", "Reyhan Khan", "All 3 days"),
    ("Producer / Ops lead", "Xarlish", "Owns execution"),
    ("Target dates", "Aug 14-16 OR Aug 28-30", "Polling Skool community this week"),
    ("Seat cap", "10", "Non-negotiable"),
    ("Base location", "Costa da Caparica, Portugal", "20 min from Lisbon"),
    ("Format", "3 days · work + experience hybrid", "Fri full · Sat full · Sun half-day content"),
    ("Pricing target", "€2,000-2,500 per seat", "Adjust based on final villa + vendor costs"),
    ("Estimated cost per seat", "€1,245-1,860", "See Budget tab"),
    ("Estimated margin per seat", "€200-700", "Accessible pricing, not margin-led"),
    ("Skool community link", "https://www.skool.com/recruitergtm", "Where seats are pitched"),
    ("Stripe checkout link", "TBC", "Set up by week 6 pre-event"),
    ("Group WhatsApp", "TBC", "Created once dates locked"),
    ("Welcome packet sent", "TBC", "4 weeks pre-event"),
    ("Reyhan's contact", "reyhan@recruitergtm.com", ""),
    ("Xarlish's contact", "TBC", ""),
]

for i, (field, value, note) in enumerate(overview_rows, start=5):
    ov.cell(row=i, column=1, value=field).font = Font(bold=True)
    ov.cell(row=i, column=2, value=value)
    ov.cell(row=i, column=3, value=note).font = Font(color="6B7280")
    for col in (1, 2, 3):
        ov.cell(row=i, column=col).border = BORDER
        ov.cell(row=i, column=col).alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    ov.row_dimensions[i].height = 24

ov.column_dimensions["A"].width = 28
ov.column_dimensions["B"].width = 38
ov.column_dimensions["C"].width = 50
ov.freeze_panes = "A5"


# ============= TAB 2: MASTER CHECKLIST =============
mc = wb.create_sheet("Master Checklist")
mc.sheet_view.showGridLines = False

mc["A1"] = "#"
mc["B1"] = "Task"
mc["C1"] = "Owner"
mc["D1"] = "Phase"
mc["E1"] = "Deadline"
mc["F1"] = "Status"
mc["G1"] = "Notes"
style_headers(mc, 1, list("ABCDEFG"))
mc.freeze_panes = "A2"

checklist = [
    # PHASE 1 — Dates + venue lock (weeks 8-7)
    ("Close community poll on dates", "Reyhan", "Phase 1 — Lock", "Week 8", "Not Started", "Pick mid-Aug 14-16 or end-Aug 28-30"),
    ("Source 5-7 villa options (10+ beds, WiFi, pool, Caparica)", "Xarlish", "Phase 1 — Lock", "Week 8", "Not Started", "Vrbo, Airbnb, Plum Guide, local rentals"),
    ("Shortlist + book primary villa", "Xarlish", "Phase 1 — Lock", "Week 7", "Not Started", "Lock by end of May"),
    ("Hold a backup villa (in case primary cancels)", "Xarlish", "Phase 1 — Lock", "Week 7", "Not Started", "Free hold if possible"),
    ("Quote 2-3 photographers + videographers (3-day shoot)", "Xarlish", "Phase 1 — Lock", "Week 7", "Not Started", "Local Lisbon talent"),
    ("Sign photographer + videographer contracts", "Reyhan", "Phase 1 — Lock", "Week 7", "Not Started", "3-day shoot + 5-day delivery"),
    ("Brief graphic designer for post-event editing", "Reyhan", "Phase 1 — Lock", "Week 7", "Not Started", "RecruiterGTM contractor"),
    # PHASE 2 — Open seats + vendor booking (weeks 6-5)
    ("Set up Stripe checkout for seats", "Reyhan", "Phase 2 — Open", "Week 6", "Not Started", "3% processing fee included"),
    ("Open waitlist + seat purchase in Skool community", "Reyhan", "Phase 2 — Open", "Week 6", "Not Started", "Announce post in Skool"),
    ("Collect first 5 deposits", "Reyhan", "Phase 2 — Open", "Week 6", "Not Started", "Track in Attendees tab"),
    ("Book catering for villa (3 breakfasts + BBQ dinner)", "Xarlish", "Phase 2 — Open", "Week 5", "Not Started", "Local Caparica catering"),
    ("Book 15-seater van + English-speaking driver (4 days)", "Xarlish", "Phase 2 — Open", "Week 5", "Not Started", "Thu-Sun"),
    ("Book surf school for 10 (Caparica)", "Xarlish", "Phase 2 — Open", "Week 5", "Not Started", "Hold both date options"),
    ("Reserve dinner at Cascais rooftop (10+ people)", "Xarlish", "Phase 2 — Open", "Week 5", "Not Started", "August high season — book NOW"),
    ("Reserve seafood lunch in Caparica beachfront", "Xarlish", "Phase 2 — Open", "Week 5", "Not Started", ""),
    ("Reserve Sunday lunch (near Cabo da Roca route)", "Xarlish", "Phase 2 — Open", "Week 5", "Not Started", ""),
    # PHASE 3 — Welcome packet + attendee info (week 4)
    ("Design + send welcome packet email to attendees", "Reyhan", "Phase 3 — Comms", "Week 4", "Not Started", "Schedule, packing, arrival info"),
    ("Build Typeform for dietary + arrival data", "Xarlish", "Phase 3 — Comms", "Week 4", "Not Started", "Collect dietary, arrival window, flight info"),
    ("Create Group WhatsApp with attendees", "Xarlish", "Phase 3 — Comms", "Week 4", "Not Started", "Pin schedule"),
    ("Visa check for non-EU attendees", "Xarlish", "Phase 3 — Comms", "Week 4", "Not Started", "Most US/UK don't need; confirm by passport"),
    # PHASE 4 — Final headcount + swag (week 3)
    ("Lock final headcount", "Reyhan", "Phase 4 — Final", "Week 3", "Not Started", "Refund deadline closes"),
    ("Order Gift Bag items (see Gift Bag tab)", "Xarlish", "Phase 4 — Final", "Week 3", "Not Started", "10-14 day production lead"),
    ("Reyhan writes 10 handwritten welcome cards", "Reyhan", "Phase 4 — Final", "Week 2", "Not Started", "Personal note per attendee"),
    # PHASE 5 — Final week (weeks 2-1)
    ("Confirm all restaurant reservations", "Xarlish", "Phase 5 — Week of", "Week 2", "Not Started", "Call to confirm"),
    ("Schedule pinned in WhatsApp", "Xarlish", "Phase 5 — Week of", "Week 1", "Not Started", ""),
    ("Equipment kit prepared (projector, cables, 4G hotspot)", "Xarlish", "Phase 5 — Week of", "Week 1", "Not Started", "Test before event"),
    ("Weather check + activity contingency confirmed", "Xarlish", "Phase 5 — Week of", "Week 1", "Not Started", "Setúbal wine or Sintra as backup"),
    # PHASE 6 — Event days
    ("Reyhan + content team arrive Thursday", "Reyhan", "Phase 6 — Event", "Day -1", "Not Started", "Villa setup, gift bag assembly"),
    ("Assemble gift bags at villa", "Xarlish + Reyhan", "Phase 6 — Event", "Day -1", "Not Started", "Insert handwritten cards"),
    ("Welcome dinner Thursday (if early arrivals)", "Reyhan", "Phase 6 — Event", "Day -1", "Not Started", "Light, optional"),
    ("Run Day 1 — Friday schedule", "Reyhan", "Phase 6 — Event", "Day 1", "Not Started", "See Schedule tab"),
    ("Run Day 2 — Saturday schedule", "Reyhan", "Phase 6 — Event", "Day 2", "Not Started", "See Schedule tab"),
    ("Run Day 3 — Sunday content + Cabo da Roca", "Reyhan", "Phase 6 — Event", "Day 3", "Not Started", "Content shoot priority"),
    ("Daily debrief with content team", "Reyhan", "Phase 6 — Event", "Daily", "Not Started", "What's working, what to adjust"),
    # PHASE 7 — Post-event
    ("Photographer delivers final content within 5 working days", "Photographer", "Phase 7 — Post", "Week +1", "Not Started", "Drive folder per attendee"),
    ("Graphic designer delivers post graphics", "Designer", "Phase 7 — Post", "Week +1", "Not Started", "Multiple per attendee"),
    ("15-min follow-up call with each attendee", "Reyhan", "Phase 7 — Post", "Week +2", "Not Started", "ATS Accelerator early-access onboarding"),
    ("Create Mastermind Alumni Slack channel", "Reyhan", "Phase 7 — Post", "Week +2", "Not Started", "Kept alive year-round"),
    ("Capture retro in year-2026-retro.md", "Reyhan", "Phase 7 — Post", "Week +3", "Not Started", "Saves to event-planner skill folder"),
    ("Reconcile actual costs vs budget", "Xarlish", "Phase 7 — Post", "Week +3", "Not Started", "Update Budget tab actuals"),
]

mc_widths = {
    "#": 5, "Task": 60, "Owner": 18, "Phase": 24,
    "Deadline": 14, "Status": 16, "Notes": 40
}
set_widths(mc, mc_widths)

for i, task in enumerate(checklist, start=2):
    mc.cell(row=i, column=1, value=i - 1).alignment = Alignment(horizontal="center", vertical="center")
    mc.cell(row=i, column=2, value=task[0])
    mc.cell(row=i, column=3, value=task[1])
    mc.cell(row=i, column=4, value=task[2])
    mc.cell(row=i, column=5, value=task[3])
    mc.cell(row=i, column=6, value=task[4])
    mc.cell(row=i, column=7, value=task[5])
    for col in range(1, 8):
        c = mc.cell(row=i, column=col)
        c.alignment = WRAP_TOP
        c.border = BORDER
        if col == 6:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(bold=True)
    mc.row_dimensions[i].height = 36

last_row = 1 + len(checklist)
add_status_dropdown(mc, "F", 2, last_row, STATUS_OPTIONS)
apply_status_conditional(mc, "F", 2, last_row, f"A2:G{last_row}")


# ============= TAB 3: BUDGET =============
bg = wb.create_sheet("Budget")
bg.sheet_view.showGridLines = False

budget_headers = ["#", "Category", "Item / Line", "Vendor", "Estimated (€)", "Actual (€)", "Status", "Notes"]
for col_idx, h in enumerate(budget_headers, start=1):
    cell = bg.cell(row=1, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = BORDER
bg.row_dimensions[1].height = 32
bg.freeze_panes = "A2"

budget_rows = [
    ("Accommodation", "Villa (4 nights — Thu-Sun)", "", 5500, "", "Researching", "Range €4,000-6,500"),
    ("Catering", "3 breakfasts at villa", "", 450, "", "Researching", "Local caterer"),
    ("Catering", "Villa BBQ dinner (Sat)", "", 350, "", "Researching", ""),
    ("Restaurants", "Cascais rooftop dinner (10+ people)", "", 800, "", "Researching", "Aug high season, book early"),
    ("Restaurants", "Caparica seafood lunch", "", 400, "", "Researching", ""),
    ("Restaurants", "Sunday lunch near Cabo da Roca", "", 300, "", "Researching", ""),
    ("Transport", "15-seater van + driver (4 days)", "", 1100, "", "Researching", "English-speaking driver"),
    ("Transport", "Airport pickup option (optional)", "", 0, "", "Researching", "Or attendees self-arrange via Uber"),
    ("Activities", "Surf class (10 people)", "", 550, "", "Researching", "Caparica surf school"),
    ("Activities", "Cristo Rei + Miradouro entries", "", 100, "", "Researching", "Per-person small fee"),
    ("Activities", "Padel court 1-hour (optional)", "", 50, "", "Researching", "If we want one"),
    ("Activities", "Cabo da Roca", "", 0, "", "Done", "Free, just transport"),
    ("Content team", "Photographer (3-day)", "", 1500, "", "Researching", ""),
    ("Content team", "Videographer (3-day)", "", 1300, "", "Researching", "Can be same person as photographer"),
    ("Content team", "Graphic designer (post-edit)", "", 700, "", "Researching", "Multiple post graphics per attendee"),
    ("Gift Bag", "RecruiterGTM tote", "Printful", 250, "", "Researching", "10 bags"),
    ("Gift Bag", "Branded notebooks (A5 hardcover)", "", 200, "", "Researching", ""),
    ("Gift Bag", "Pens with logo", "", 50, "", "Researching", ""),
    ("Gift Bag", "Sticker pack", "StickerMule", 80, "", "Researching", "RecruiterGTM + Claude themed"),
    ("Gift Bag", "Custom Lisbon postcards", "", 60, "", "Researching", "Local designer + Reyhan handwrites"),
    ("Gift Bag", "Reusable water bottles", "", 150, "", "Researching", ""),
    ("Gift Bag", "Sunscreen + lip balm", "", 80, "", "Researching", ""),
    ("Gift Bag", "Pastéis de Nata vouchers", "Manteigaria", 50, "", "Researching", ""),
    ("Gift Bag", "Schedule card (printed)", "", 30, "", "Researching", ""),
    ("Gift Bag", "Optional: Portuguese wine bottle", "", 200, "", "Researching", "Premium add-on"),
    ("Gift Bag", "Optional: Lisbon art print", "", 250, "", "Researching", "Annual limited edition"),
    ("Misc", "Drinks at villa (wine, beer, water, coffee)", "", 350, "", "Researching", ""),
    ("Misc", "Equipment (projector, cables, 4G hotspot)", "", 250, "", "Researching", ""),
    ("Misc", "Stationery (pens, A4 notepads on the day)", "", 100, "", "Researching", ""),
    ("Misc", "Insurance — public liability for 3 days", "", 200, "", "Researching", "Check villa requirements"),
    ("Misc", "Stripe processing (3% of revenue)", "", 600, "", "Researching", "Estimated on €20k revenue"),
    ("Misc", "Buffer 15%", "", 1500, "", "Researching", "Always include"),
]

for i, row in enumerate(budget_rows, start=2):
    bg.cell(row=i, column=1, value=i - 1).alignment = Alignment(horizontal="center", vertical="center")
    bg.cell(row=i, column=2, value=row[0])
    bg.cell(row=i, column=3, value=row[1])
    bg.cell(row=i, column=4, value=row[2])
    bg.cell(row=i, column=5, value=row[3]).number_format = "#,##0 €"
    bg.cell(row=i, column=6, value=row[4]).number_format = "#,##0 €"
    bg.cell(row=i, column=7, value=row[5])
    bg.cell(row=i, column=8, value=row[6])
    for col in range(1, 9):
        c = bg.cell(row=i, column=col)
        c.alignment = WRAP_TOP
        c.border = BORDER
        if col == 7:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(bold=True)
    bg.row_dimensions[i].height = 28

# Total row
total_row = len(budget_rows) + 2
bg.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True, size=12)
bg.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row-1})").number_format = "#,##0 €"
bg.cell(row=total_row, column=5).font = Font(bold=True, color="8A00FF")
bg.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row-1})").number_format = "#,##0 €"
bg.cell(row=total_row, column=6).font = Font(bold=True, color="8A00FF")
for col in range(1, 9):
    bg.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="F3E8FF")
    bg.cell(row=total_row, column=col).border = BORDER

# Per-seat row
seat_row = total_row + 1
bg.cell(row=seat_row, column=2, value="Per seat (÷10)").font = Font(bold=True)
bg.cell(row=seat_row, column=5, value=f"=E{total_row}/10").number_format = "#,##0 €"
bg.cell(row=seat_row, column=5).font = Font(bold=True)
bg.cell(row=seat_row, column=6, value=f"=F{total_row}/10").number_format = "#,##0 €"
bg.cell(row=seat_row, column=6).font = Font(bold=True)

set_widths(bg, {"#": 5, "Category": 18, "Item / Line": 38, "Vendor": 22, "Est": 14, "Act": 14, "Status": 16, "Notes": 36})
add_status_dropdown(bg, "G", 2, len(budget_rows) + 1, VENDOR_STATUS)


# ============= TAB 4: VILLA SHORTLIST =============
vl = wb.create_sheet("Villa Shortlist")
vl.sheet_view.showGridLines = False

villa_headers = ["#", "Villa name", "URL / source", "Beds", "Price / night (€)", "Total 4 nights (€)", "WiFi check", "Pool / Beach", "Parking", "Status", "Notes"]
for col_idx, h in enumerate(villa_headers, start=1):
    cell = vl.cell(row=1, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = BORDER
vl.row_dimensions[1].height = 32
vl.freeze_panes = "A2"

# 7 empty rows for Xarlish to fill
for i in range(2, 9):
    vl.cell(row=i, column=1, value=i - 1).alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 12):
        c = vl.cell(row=i, column=col)
        c.border = BORDER
        c.alignment = WRAP_TOP
    vl.row_dimensions[i].height = 36

set_widths(vl, {"#": 5, "Villa": 28, "URL": 36, "Beds": 8, "Price/night": 16, "Total": 18, "WiFi": 22, "Pool": 18, "Parking": 14, "Status": 16, "Notes": 36})
add_status_dropdown(vl, "J", 2, 8, ["Researching", "Visited", "Shortlisted", "Primary", "Backup", "Rejected"])


# ============= TAB 5: VENDORS =============
vd = wb.create_sheet("Vendors")
vd.sheet_view.showGridLines = False

vendor_headers = ["#", "Category", "Vendor name", "Contact", "Quote (€)", "Status", "Notes"]
for col_idx, h in enumerate(vendor_headers, start=1):
    cell = vd.cell(row=1, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = BORDER
vd.row_dimensions[1].height = 32
vd.freeze_panes = "A2"

vendor_rows = [
    ("Photographer", "", "", 1500, "Researching", "3-day shoot, 5-day delivery"),
    ("Videographer", "", "", 1300, "Researching", "Can be same person, ask for combo rate"),
    ("Graphic designer", "", "", 700, "Researching", "Post-event editing"),
    ("Catering — villa", "", "", 800, "Researching", "Breakfasts + BBQ"),
    ("Van + driver", "", "", 1100, "Researching", "15-seater, 4 days, English-speaking"),
    ("Surf school", "", "", 550, "Researching", "Caparica, block both date options"),
    ("Restaurant — Cascais rooftop", "", "", 800, "Researching", "Book NOW for August"),
    ("Restaurant — Caparica seafood", "", "", 400, "Researching", ""),
    ("Restaurant — Sunday lunch", "", "", 300, "Researching", ""),
    ("Padel court (optional)", "", "", 50, "Researching", "1-hour booking"),
    ("Gift bag — tote + apparel", "Printful", "", 250, "Researching", ""),
    ("Gift bag — stickers", "StickerMule", "", 80, "Researching", ""),
    ("Gift bag — notebooks", "", "", 200, "Researching", "Custom branded A5 hardcover"),
    ("Gift bag — postcards", "", "", 60, "Researching", "Local Lisbon designer"),
    ("Gift bag — wine (optional)", "", "", 200, "Researching", "Local Lisbon wine shop"),
    ("Insurance", "", "", 200, "Researching", "Public liability for 3 days"),
]

for i, row in enumerate(vendor_rows, start=2):
    vd.cell(row=i, column=1, value=i - 1).alignment = Alignment(horizontal="center", vertical="center")
    vd.cell(row=i, column=2, value=row[0])
    vd.cell(row=i, column=3, value=row[1])
    vd.cell(row=i, column=4, value=row[2])
    vd.cell(row=i, column=5, value=row[3]).number_format = "#,##0 €"
    vd.cell(row=i, column=6, value=row[4])
    vd.cell(row=i, column=7, value=row[5])
    for col in range(1, 8):
        c = vd.cell(row=i, column=col)
        c.alignment = WRAP_TOP
        c.border = BORDER
        if col == 6:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(bold=True)
    vd.row_dimensions[i].height = 32

set_widths(vd, {"#": 5, "Category": 28, "Vendor": 24, "Contact": 28, "Quote": 14, "Status": 16, "Notes": 36})
add_status_dropdown(vd, "F", 2, len(vendor_rows) + 1, VENDOR_STATUS)


# ============= TAB 6: ATTENDEES =============
at = wb.create_sheet("Attendees")
at.sheet_view.showGridLines = False

attendee_headers = ["#", "Name", "Company", "Email", "Country", "Arrival", "Departure", "Dietary", "Seat (€)", "Payment", "Notes"]
for col_idx, h in enumerate(attendee_headers, start=1):
    cell = at.cell(row=1, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = BORDER
at.row_dimensions[1].height = 32
at.freeze_panes = "A2"

# 10 empty rows for attendees
for i in range(2, 12):
    at.cell(row=i, column=1, value=i - 1).alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 12):
        c = at.cell(row=i, column=col)
        c.border = BORDER
        c.alignment = WRAP_TOP
    at.row_dimensions[i].height = 32

set_widths(at, {"#": 5, "Name": 22, "Company": 24, "Email": 28, "Country": 14, "Arrival": 14, "Departure": 14, "Dietary": 20, "Seat": 12, "Payment": 18, "Notes": 30})
add_status_dropdown(at, "J", 2, 11, PAYMENT_OPTIONS)
# Color rows by payment status
at.conditional_formatting.add(
    "A2:K11",
    FormulaRule(formula=['$J2="Paid in Full"'], fill=PAID_FILL),
)
at.conditional_formatting.add(
    "A2:K11",
    FormulaRule(formula=['$J2="Deposit Paid"'], fill=PENDING_FILL),
)


# ============= TAB 7: SCHEDULE =============
sc = wb.create_sheet("Schedule")
sc.sheet_view.showGridLines = False

sc_headers = ["Day", "Time", "Activity", "Location", "Lead", "Notes"]
for col_idx, h in enumerate(sc_headers, start=1):
    cell = sc.cell(row=1, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = BORDER
sc.row_dimensions[1].height = 32
sc.freeze_panes = "A2"

schedule = [
    ("Thursday (Day -1)", "PM", "Reyhan + content team arrive, villa setup", "Villa", "Reyhan + Xarlish", "Equipment check, gift bag assembly"),
    ("Thursday (Day -1)", "Evening", "Welcome dinner (early arrivals, optional)", "Local Caparica restaurant", "Reyhan", "Light"),
    ("Friday (Day 1)", "09:00-12:00", "Group breakfast + welcome session", "Villa common area", "Reyhan", "Intro round, set personal 3-day goals"),
    ("Friday (Day 1)", "12:00-14:00", "ATS Accelerator workshop sprint #1", "Villa common area", "Reyhan", "Each attendee starts their ATS map"),
    ("Friday (Day 1)", "14:00-17:00", "Cristo Rei + Lisbon Miradouro tour", "Across river", "Xarlish + van driver", "Photo opportunities"),
    ("Friday (Day 1)", "17:00-19:00", "Beach + free time", "Caparica beach", "Free", ""),
    ("Friday (Day 1)", "19:00-22:00", "Dinner at Cascais rooftop", "Cascais", "Xarlish", "Reservation booked Week 5"),
    ("Saturday (Day 2)", "09:00-12:00", "Claude AI installation sprint (1-on-1)", "Villa", "Reyhan", "Each attendee gets their Claude Ops Manager wired"),
    ("Saturday (Day 2)", "12:00-14:00", "Seafood lunch", "Caparica beachfront", "Xarlish", "Reservation booked Week 5"),
    ("Saturday (Day 2)", "14:00-17:00", "Surf class", "Caparica beach", "Surf school", "All 10 attendees"),
    ("Saturday (Day 2)", "17:00-19:00", "Free time / cocktails", "Villa", "Free", ""),
    ("Saturday (Day 2)", "19:00-22:00", "BBQ at villa", "Villa terrace", "Caterer", "Casual founder dinner"),
    ("Sunday (Day 3)", "09:00-12:00", "BD campaign launch sprint + LinkedIn reel filming", "Villa", "Reyhan + photographer", "1 reel per attendee"),
    ("Sunday (Day 3)", "12:00-14:00", "Lunch + drive to Cabo da Roca", "Sintra route", "Xarlish", ""),
    ("Sunday (Day 3)", "14:00", "Cabo da Roca", "Westernmost Europe", "Reyhan", "Photo + sunset (depending on date)"),
    ("Sunday (Day 3)", "PM", "Casual goodbye, attendees depart", "Villa / airport", "Xarlish", "Coordinate airport runs"),
]

for i, row in enumerate(schedule, start=2):
    for col_idx, val in enumerate(row, start=1):
        c = sc.cell(row=i, column=col_idx, value=val)
        c.alignment = WRAP_TOP
        c.border = BORDER
    sc.row_dimensions[i].height = 36

set_widths(sc, {"Day": 18, "Time": 14, "Activity": 42, "Location": 24, "Lead": 18, "Notes": 36})


# ============= TAB 8: GIFT BAG =============
gb = wb.create_sheet("Gift Bag")
gb.sheet_view.showGridLines = False

gb_headers = ["#", "Item", "Vendor", "Qty", "Cost / unit (€)", "Total (€)", "Lead time", "Status", "Notes"]
for col_idx, h in enumerate(gb_headers, start=1):
    cell = gb.cell(row=1, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = BORDER
gb.row_dimensions[1].height = 32
gb.freeze_panes = "A2"

gift_bag = [
    ("Custom RecruiterGTM tote bag", "Printful / local screen printer", 10, 25, "=D2*E2", "10-14 days", "Researching", "Carries the rest"),
    ("A5 hardcover notebook (branded)", "Moleskine custom OR local stationer", 10, 20, "=D3*E3", "14 days", "Researching", ""),
    ("Pen with logo", "Vistaprint or local", 10, 5, "=D4*E4", "7 days", "Researching", ""),
    ("Sticker pack (RecruiterGTM + Claude)", "StickerMule", 10, 8, "=D5*E5", "7 days", "Researching", ""),
    ("Custom Lisbon postcard (designed)", "Local Lisbon designer", 10, 6, "=D6*E6", "10 days", "Researching", "Reyhan handwrites each one"),
    ("Reusable water bottle (logo)", "", 10, 15, "=D7*E7", "10 days", "Researching", "Useful all 3 days"),
    ("Sunscreen + lip balm", "", 10, 8, "=D8*E8", "7 days", "Researching", "Lisbon summer essential"),
    ("Pastel de Nata voucher", "Manteigaria / Pastéis de Belém", 10, 5, "=D9*E9", "7 days", "Researching", "Pre-paid voucher"),
    ("Schedule card (A6 printed)", "Local print shop", 10, 3, "=D10*E10", "5 days", "Researching", ""),
    ("Bottle of Portuguese wine (optional)", "Local wine shop", 10, 20, "=D11*E11", "On the day", "Researching", "Premium add-on"),
    ("Lisbon art print limited edition (optional)", "Local artist commission", 10, 25, "=D12*E12", "21 days", "Researching", "Year 1 tradition starter"),
]

for i, row in enumerate(gift_bag, start=2):
    gb.cell(row=i, column=1, value=i - 1).alignment = Alignment(horizontal="center", vertical="center")
    gb.cell(row=i, column=2, value=row[0])
    gb.cell(row=i, column=3, value=row[1])
    gb.cell(row=i, column=4, value=row[2]).alignment = Alignment(horizontal="center", vertical="center")
    gb.cell(row=i, column=5, value=row[3]).number_format = "#,##0 €"
    gb.cell(row=i, column=6, value=row[4]).number_format = "#,##0 €"
    gb.cell(row=i, column=7, value=row[5]).alignment = Alignment(horizontal="center", vertical="center")
    gb.cell(row=i, column=8, value=row[6])
    gb.cell(row=i, column=9, value=row[7])
    for col in range(1, 10):
        c = gb.cell(row=i, column=col)
        c.border = BORDER
        if c.alignment.horizontal not in ("center",):
            c.alignment = WRAP_TOP
        if col == 8:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(bold=True)
    gb.row_dimensions[i].height = 32

# Total
gb_total = len(gift_bag) + 2
gb.cell(row=gb_total, column=2, value="GIFT BAG TOTAL").font = Font(bold=True, size=12)
gb.cell(row=gb_total, column=6, value=f"=SUM(F2:F{gb_total-1})").number_format = "#,##0 €"
gb.cell(row=gb_total, column=6).font = Font(bold=True, color="8A00FF")
gb.cell(row=gb_total + 1, column=2, value="Per attendee (÷10)").font = Font(bold=True)
gb.cell(row=gb_total + 1, column=6, value=f"=F{gb_total}/10").number_format = "#,##0 €"
gb.cell(row=gb_total + 1, column=6).font = Font(bold=True)
for col in range(1, 10):
    gb.cell(row=gb_total, column=col).fill = PatternFill("solid", fgColor="F3E8FF")
    gb.cell(row=gb_total, column=col).border = BORDER

set_widths(gb, {"#": 5, "Item": 36, "Vendor": 24, "Qty": 8, "Cost/unit": 14, "Total": 14, "Lead": 14, "Status": 16, "Notes": 28})
add_status_dropdown(gb, "H", 2, len(gift_bag) + 1, VENDOR_STATUS)


# ─── Save ──────────────────────────────────────────────────────────────
out_path = Path(os.path.expanduser("~/Desktop/lisbon-mastermind-planner-2026.xlsx"))
wb.save(out_path)
print(f"Saved: {out_path}")
print("  Tabs: Overview · Master Checklist · Budget · Villa Shortlist · Vendors · Attendees · Schedule · Gift Bag")
