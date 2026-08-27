#!/usr/bin/env python3
"""RecruiterGTM 90-Day YouTube Tracker (.xlsx) - audience-optimized, funnel-mapped, editor-shareable."""
import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/reyhankhan/Library/CloudStorage/GoogleDrive-reyhan@recruitergtm.com/My Drive/EA Demo/projects/youtube-channel-launch/RecruiterGTM YouTube 90-Day Tracker.xlsx"

VIOLET = "8A00FF"; DARK = "1A1A1A"; LIGHT = "F2E9FF"; CREAM = "FFF4E6"
TECH = "AI & Tech for Recruiters"; FOUND = "Founder Lessons"

start = dt.date(2026, 8, 12)
while start.weekday() != 1:  # first Tuesday on/after
    start += dt.timedelta(days=1)

# Title, Series, RGTM Pillar/Offer, Format, Primary Keyword (SEO), Offer/Funnel (CTA + what it feeds), EditStyle, ThumbText
rows = [
    ("How to Install a Claude AI Ops Manager in Your Recruitment Agency", TECH,
     "Pillar 1: AI Layer - Skool + Claude-Code-DFY", "Tutorial", "claude ai for recruitment agency",
     "Skool value-prop #1 (Claude Ops Manager). CTA: join Skool. Proof: 650+ agencies audited.", "Screen-share", "AI OPS MANAGER"),
    ("The 3-Engine GTM System Every Recruitment Agency Needs", TECH,
     "5 Pillars overview - 3 DFY Pilots (BD/Sourcing/Content)", "Framework", "gtm system for recruitment agency",
     "Top-funnel for the 3 DFY pilots. CTA: Book a call.", "Screen-share", "3 ENGINES"),
    ("Clay for Recruiters: Build a Live Jobs Lead List in 10 Minutes", TECH,
     "Pillar 2: Multichannel Outbound - BD Pilot", "Tutorial", "clay for recruiters",
     "BD pilot demo. Clay affiliate link. CTA: Skool / newsletter.", "Screen-share", "CLAY x JOBS"),
    ("3 Ways Recruitment Agencies Should Be Using AI in 2026", TECH,
     "Pillar 1: AI Layer - Claude Code", "Framework", "ai for recruitment agencies 2026",
     "Broad top-funnel. CTA: claude-code-dfy page.", "Screen-share", "AI x3"),
    ("Build a $600/Month Outbound Engine for Your Agency (Full Stack)", TECH,
     "Pillar 2: Multichannel Outbound - BD Pilot", "Tutorial", "recruitment outbound tech stack",
     "Instantly + HeyReach + Clay + Apollo (all affiliate). CTA: outbound-os / book a call.", "Screen-share", "$600 STACK"),
    ("SourcingOS: An AI Candidate Sourcing Machine for Any Role", TECH,
     "Pillar 4: ATS + Database - Sourcing Pilot", "Tutorial", "ai candidate sourcing",
     "Sourcing pilot demo (Stardex ATS). CTA: Book a call.", "Screen-share", "SOURCING OS"),
    ("How I Built a LinkedIn Content Engine with Claude (ContentOS)", TECH,
     "Pillar 3: Content + Authority - Content Pilot", "Framework", "linkedin content system for recruiters",
     "Content pilot + Beehiiv newsletter nurture. CTA: newsletter / Skool.", "Screen-share", "CONTENT ENGINE"),
    ("HeyReach for Recruiters: LinkedIn Outreach That Books Calls", TECH,
     "Pillar 2: Multichannel Outbound - BD Pilot", "Tutorial", "heyreach linkedin outreach",
     "BD pilot. HeyReach affiliate. CTA: outbound-os.", "Screen-share", "HEYREACH"),
    ("How to Hire an Offshore GTM Engineer (What They Actually Do)", TECH,
     "Pillar 5: Productization - Talent Placement (public)", "Framework", "hire offshore gtm engineer",
     "Talent offer (Noroze). CTA: Book a call - sample candidates + video intros shown at call stage. Proof: 200+ placed, 98% at 18mo.", "Screen-share", "THE RIGHT HIRE"),
    ("$0 to $12k/Month as a Recruitment Ops Expert (My Story)", FOUND,
     "Founder story - Positioning", "Story", "recruitment operations expert",
     "Authority + newsletter nurture. CTA: newsletter.", "Cinematic", "$0 to $12K"),
    ("Master of None: Why 'Focus on One Thing' Was Bad Advice", FOUND,
     "Founder / contrarian", "Contrarian", "generalist vs specialist",
     "Relatability + subscribe. CTA: newsletter.", "Cinematic", "MASTER OF NONE"),
    ("0 to $20k in 20 Days: How I Niched into Recruitment GTM", FOUND,
     "Founder story - Niche down", "Story", "how to niche your agency",
     "Origin story -> the offer. CTA: newsletter / Skool.", "Cinematic", "20 DAYS"),
    ("Why 80% of Agency Burnout Is Ops, Not Sales", FOUND,
     "OperatorOS - Diagnostic", "Contrarian", "recruitment agency burnout",
     "Pain -> the AI Ops fix. CTA: Book a call.", "Cinematic", "BURNOUT"),
]

# Title, Series, Pillar/Offer, Funnel
reserve = [
    ("AI Agents for Recruiters, Explained (No Code)", TECH, "Pillar 1: AI Layer", "claude-code-dfy"),
    ("Rank Your Recruitment Agency in AI Search (AI SEO)", TECH, "Skool value-prop #3: Website/SEO", "seo-engine / Skool"),
    ("The Biggest Operational Gaps in a Recruitment Agency", TECH, "OperatorOS", "Book a call"),
    ("3 Clay Playbooks Under 10 Min (Funding, Talent Replacement)", TECH, "Pillar 2/4 - BD + Sourcing", "Clay affiliate / Skool"),
    ("Top 3 n8n Automations to 3x Agency Productivity", TECH, "Pillar 1: AI Layer", "claude-code-dfy"),
    ("How AI Lifts Agency Profit Margins with a Lean Team", TECH, "Pillar 5: Productization", "Book a call"),
    ("Train Your Claude to Make You 4x More Productive", TECH, "Pillar 1: AI Layer", "Skool"),
    ("Quick-Wins Campaign: Revenue From Your LinkedIn Network", TECH, "Pillar 2: Outbound", "newsletter"),
    ("One Hack to Land More Placements: The Video Intro", TECH, "Talent Placement", "Book a call"),
    ("The Remote Work Blueprint (How I Went Fully Remote)", FOUND, "Founder story", "newsletter"),
    ("The AI Adoption Arc: Get Ahead Before Everyone Else", FOUND, "Contrarian / AI", "claude-code-dfy"),
    ("How to Find Your Next High-Ticket Recruitment Client", FOUND, "Sales / positioning", "Book a call"),
]

# Short idea, Series, Funnel angle
shorts_bank = [
    ("Clip: the Claude Ops Manager doing one real task live", TECH, "claude-code-dfy"),
    ("AI won't replace recruiters who do THIS", TECH, "claude-code-dfy"),
    ("Clay: live jobs list in 5 minutes", TECH, "Skool"),
    ("How I make LinkedIn content on the go", TECH, "newsletter"),
    ("The IKEA analogy - I build it WITH recruiters, not FOR them", FOUND, "Book a call"),
    ("Life is a video game", FOUND, "subscribe"),
    ("The single greatest skill: be in a good mood for no reason", FOUND, "subscribe"),
    ("20 Things I Wish I Knew at 20 (series)", FOUND, "newsletter"),
    ("The education system makes you employable, not rich", FOUND, "subscribe"),
    ("How I got into the top 0.5% in the world", FOUND, "newsletter"),
    ("Manifestation in my faith (personal / real)", FOUND, "subscribe"),
    ("How I use systems in my personal life to save time", FOUND, "newsletter"),
    ("Monetize your hobbies", FOUND, "subscribe"),
]

thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
head_fill = PatternFill("solid", fgColor=DARK)
head_font = Font(bold=True, color="FFFFFF", size=11)


def banner(ws, text, span, sub=None):
    ws.merge_cells(f"A1:{span}1")
    c = ws["A1"]; c.value = text
    c.font = Font(size=15, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=VIOLET)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    if sub:
        ws.merge_cells(f"A2:{span}2")
        s = ws["A2"]; s.value = sub
        s.font = Font(size=10, italic=True, color="666666")
        s.alignment = Alignment(horizontal="left", indent=1)
        ws.row_dimensions[2].height = 28


wb = openpyxl.Workbook()

# ---------- Sheet 1: 90-Day Slate ----------
ws = wb.active
ws.title = "90-Day Slate"
banner(ws, "RecruiterGTM - 90-Day YouTube Production Tracker", "M",
       "Audience: recruitment agency owners/founders + GTM engineers. Every video is top-of-funnel for a RecruiterGTM offer. "
       "Colour = series (violet: AI & Tech / cream: Founder Lessons). See Strategy + Series Guide tabs.")

headers = ["Wk", "Target Publish", "Series", "Title", "RGTM Pillar / Offer", "Format",
           "Primary Keyword (SEO)", "Offer / Funnel", "Edit Style", "Status", "Thumbnail Text", "Editor", "Editor Notes"]
hdr_row = 4
for j, h in enumerate(headers, 1):
    cell = ws.cell(row=hdr_row, column=j, value=h)
    cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[hdr_row].height = 30

for i, (t, series, pillar, fmt, kw, funnel, style, thumb) in enumerate(rows):
    r = hdr_row + 1 + i
    pub = start + dt.timedelta(weeks=i)
    vals = [i + 1, pub.strftime("%a %d %b %Y"), series, t, pillar, fmt, kw, funnel, style, "Idea", thumb, "Shayan", ""]
    tint = LIGHT if series == TECH else CREAM
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor=tint)
        cell.alignment = Alignment(vertical="center", wrap_text=(j in (4, 5, 7, 8, 13)),
                                   horizontal="center" if j in (1, 10) else "left",
                                   indent=0 if j in (1, 10) else 1)

dv = DataValidation(type="list", formula1='"Idea,Scripting,Filming,Editing,Thumbnail,Scheduled,Published"', allow_blank=True)
ws.add_data_validation(dv); dv.add(f"J{hdr_row+1}:J{hdr_row+len(rows)}")
widths = [4, 15, 22, 40, 30, 11, 24, 40, 12, 11, 15, 9, 22]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A5"

# ---------- Sheet 2: Strategy ----------
wss = wb.create_sheet("Strategy")
banner(wss, "Channel Strategy - how every video feeds RecruiterGTM", "B")
def sec(ws, r, title):
    ws.merge_cells(f"A{r}:B{r}")
    c = ws.cell(row=r, column=1, value=title)
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=VIOLET)
    c.alignment = Alignment(indent=1, vertical="center"); ws.row_dimensions[r].height = 20
def kv(ws, r, k, v):
    a = ws.cell(row=r, column=1, value=k); a.font = Font(bold=True); a.alignment = Alignment(vertical="top", indent=1, wrap_text=True)
    b = ws.cell(row=r, column=2, value=v); b.alignment = Alignment(vertical="top", wrap_text=True, indent=1)

r = 4
sec(wss, r, "North Star"); r += 1
kv(wss, r, "Mission", "Make Reyhan the #1 Claude + AI systems implementer for recruitment agencies. The channel is top-of-funnel for that positioning."); r += 1
kv(wss, r, "Audience", "Recruitment agency owners & founders (primary); GTM engineers / ops hires (secondary)."); r += 1
kv(wss, r, "Flagship angle", "Lead with Claude / AI implementation. Talent placement is now public (Noroze) and can be featured alongside."); r += 2

sec(wss, r, "The 5 Pillars (tag every video to one)"); r += 1
for k, v in [
    ("1. AI Layer", "Custom Claude Ops Manager + skills + MCP to their stack."),
    ("2. Multichannel Outbound", "Email + LinkedIn + dialer. Intent-based playbooks (Instantly, HeyReach, Clay, Apollo)."),
    ("3. Content + Authority", "LinkedIn + YouTube + Beehiiv newsletter around the operator."),
    ("4. ATS + Database", "Clean ATS (Stardex) as source of truth, segmentation, re-engagement."),
    ("5. Productization", "Offer structure, pricing, retained/fractional tiers.")]:
    kv(wss, r, k, v); r += 1
r += 1

sec(wss, r, "Offers the channel funnels to"); r += 1
for k, v in [
    ("DFY 90-Day Pilots (primary)", "One engine each: Business Development / Sourcing / Content. Benchmark ~$2,500/mo. CTA: Book a call."),
    ("Skool Community", "$1,497 one-time, 12 mo. 3 value props: Claude Ops Manager + 1st Intent-Based BD Campaign + Website/SEO."),
    ("Talent Placement (public)", "Offshore GTM engineers / ops / recruiters. Noroze owns delivery. CTA: Book a call."),
    ("Retainers", "~$2k/mo GTM engine management. CTA: Book a call."),
    ("Affiliates", "Clay, Instantly, HeyReach, Apollo, Apify, Pin.com - always use affiliate links in descriptions.")]:
    kv(wss, r, k, v); r += 1
r += 1

sec(wss, r, "Funnel logic"); r += 1
for k, v in [
    ("AI & Tech videos", "Value-first demo -> mid-roll CTA to Skool / claude-code-dfy / outbound-os / book a call. Tool mentions use affiliate links."),
    ("Founder Lessons videos", "Story / mindset -> soft CTA to newsletter + subscribe. Builds trust, not a hard sell."),
    ("Promo cap", "Keep promotional videos <=30% of the mix. This slate is mostly value - good."),
    ("Session time", "Every description ends with a 'Watch next' link to a related video.")]:
    kv(wss, r, k, v); r += 1
r += 1

sec(wss, r, "Guardrails (verify before recording)"); r += 1
for k, v in [
    ("No revenue promises", "We install the 5-pillar system. Sanctioned outcomes: profitability up, stress down, more conversations. Client numbers = testimonials only."),
    ("Talent placement", "Never pitch without >=2 real sample candidates + video intros - show these at the call/booking stage, not in the video."),
    ("Entity", "RecruiterGTM LLC on any on-screen legal/branding."),
    ("Approved stats", "650+ agencies audited; 200+ GTME placements (98% still in role at 18mo); 50+ OutboundOS deployments; 82 Skool members; 4 yrs recruitment systems / 8 yrs GTM. NOT valid: '250+ placed across 300+ companies'."),
    ("Prices", "Don't invent tool prices; keep hard prices out of titles.")]:
    kv(wss, r, k, v); r += 1
for rr in range(4, r):
    for j in range(1, 3):
        wss.cell(row=rr, column=j).border = border
wss.column_dimensions["A"].width = 26
wss.column_dimensions["B"].width = 92

# ---------- Sheet 3: Series Guide ----------
wsg = wb.create_sheet("Series Guide")
banner(wsg, "Series Guide - for Shayan (editor)", "C")
guide = [
    ("Series", TECH, FOUND),
    ("What it is", "Tool feedback, tutorials, AI/Claude use cases, system walkthroughs.", "Reyhan's story, mindset, lifestyle, contrarian takes."),
    ("Tone", "Practical, fast, authoritative. Practitioner not consultant.", "Personal, reflective, real. Honest about being early-stage."),
    ("Edit style", "Screen-share heavy. Zoom-ins on clicks, dashboard b-roll, snappy cuts, burned-in captions, chapter markers.", "Cinematic. Location / drone / lifestyle b-roll, slower pacing, music-led, face-to-camera, minimal text."),
    ("Pacing", "Fast - cut dead air hard.", "Breathe - let moments land."),
    ("Thumbnail", "Dark bg + violet #8A00FF glow, tool logo / dashboard, 1-2 words.", "Face-forward, location backdrop, violet accent, 1-2 words."),
    ("Captions", "Always - burned-in, keyword-styled.", "Light / optional - don't clutter."),
    ("CTA", "Mid-roll to Skool / claude-code-dfy / outbound-os + end screen.", "Soft - newsletter + subscribe."),
    ("Length target", "6-12 min.", "5-9 min."),
]
for i, (label, a, b) in enumerate(guide):
    r = 4 + i
    lab = wsg.cell(row=r, column=1, value=label); lab.font = Font(bold=True, color="FFFFFF")
    lab.fill = PatternFill("solid", fgColor=DARK); lab.alignment = Alignment(vertical="top", indent=1)
    ca = wsg.cell(row=r, column=2, value=a); cb = wsg.cell(row=r, column=3, value=b)
    ca.fill = PatternFill("solid", fgColor=LIGHT); cb.fill = PatternFill("solid", fgColor=CREAM)
    for cc in (ca, cb): cc.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    for j in range(1, 4): wsg.cell(row=r, column=j).border = border
wsg.column_dimensions["A"].width = 16; wsg.column_dimensions["B"].width = 52; wsg.column_dimensions["C"].width = 52
sr = 4 + len(guide) + 1
wsg.cell(row=sr, column=1, value="Deliverables per video (both series):").font = Font(bold=True, italic=True)
for i, s in enumerate([
    "Long-form master (16:9, 1080p+), colour-corrected, audio levelled.",
    "3-5 vertical shorts (9:16) from the best moments - see Shorts Bank tab.",
    "Thumbnail: 1280x720, readable at 10% zoom, 1-2 words max.",
    "File naming: [Wk##]-[Series]-[Short Title].",
    "Source footage / scripts live in the Drive YouTube folder (same place as this tracker)."]):
    wsg.cell(row=sr + 1 + i, column=1, value="- " + s)

# ---------- Sheet 4: Reserve ----------
ws2 = wb.create_sheet("Reserve Topics")
banner(ws2, "Reserve Topics - swap in if a week isn't ready", "D")
for j, h in enumerate(["#", "Title", "Series", "RGTM Pillar / Offer", "Funnel"], 1):
    cell = ws2.cell(row=3, column=j, value=h); cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal="center")
for i, (t, series, pillar, funnel) in enumerate(reserve):
    r = 4 + i
    for j, v in enumerate([i + 1, t, series, pillar, funnel], 1):
        cell = ws2.cell(row=r, column=j, value=v)
        cell.fill = PatternFill("solid", fgColor=(LIGHT if series == TECH else CREAM)); cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(j in (2, 4)), indent=1, horizontal="center" if j == 1 else "left")
for col, w in zip("ABCDE", [4, 46, 22, 30, 18]): ws2.column_dimensions[col].width = w

# ---------- Sheet 5: Shorts Bank ----------
ws3 = wb.create_sheet("Shorts Bank")
banner(ws3, "Shorts Bank - 5/week from long-form clips + these standalone hooks", "D")
for j, h in enumerate(["#", "Short Idea", "Series", "Funnel", "Status"], 1):
    cell = ws3.cell(row=3, column=j, value=h); cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal="center")
dv3 = DataValidation(type="list", formula1='"Idea,Filming,Editing,Scheduled,Published"', allow_blank=True)
ws3.add_data_validation(dv3)
for i, (t, series, funnel) in enumerate(shorts_bank):
    r = 4 + i
    for j, v in enumerate([i + 1, t, series, funnel, "Idea"], 1):
        cell = ws3.cell(row=r, column=j, value=v)
        cell.fill = PatternFill("solid", fgColor=(LIGHT if series == TECH else CREAM)); cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(j == 2), indent=1, horizontal="center" if j == 1 else "left")
dv3.add(f"E4:E{3+len(shorts_bank)}")
for col, w in zip("ABCDE", [4, 52, 22, 16, 12]): ws3.column_dimensions[col].width = w

wb.save(OUT)
print("Saved:", OUT)
print("First publish:", start.strftime("%a %d %b %Y"))
print("Tabs: 90-Day Slate | Strategy | Series Guide | Reserve Topics | Shorts Bank")
